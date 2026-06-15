#!/usr/bin/env python3
"""Export team skill register rows from Excel into a web-friendly JSON file."""

from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = ROOT / "shared-templates" / "team_skill_register.xlsx"
DEFAULT_OUTPUT = ROOT / "data" / "team_skill_register.json"
REGISTER_SHEET = "已制作Skill登记"


def normalize(value):
    if value is None:
        return ""
    return str(value).strip()


def row_to_item(row):
    return {
        "Skill_ID": normalize(row.get("Skill_ID")),
        "L2_ID": normalize(row.get("Skill_ID")),
        "L2 Skill名称": normalize(row.get("Skill名称")),
        "场景大类": normalize(row.get("场景大类")) or "客户金融场景",
        "行业": normalize(row.get("行业")),
        "所属层级": normalize(row.get("所属层级")),
        "业务子类": normalize(row.get("业务子类")),
        "业务环节": normalize(row.get("业务环节")),
        "能力说明": normalize(row.get("能力说明")),
        "Skill分类": normalize(row.get("Skill分类")),
        "关键输入": normalize(row.get("关键输入")),
        "关键输出": normalize(row.get("关键输出")),
        "建设状态": normalize(row.get("建设状态")),
        "展示状态": normalize(row.get("展示状态")),
        "作者": normalize(row.get("作者")),
        "来源平台": normalize(row.get("来源平台")),
        "迁移建议": normalize(row.get("迁移建议")),
        "Skill文件路径": normalize(row.get("Skill文件路径")),
        "Demo页面路径": normalize(row.get("Demo页面路径")),
        "规则/模板路径": normalize(row.get("规则/模板路径")),
        "样例数据路径": normalize(row.get("样例数据路径")),
        "版本": normalize(row.get("版本")),
        "标签": normalize(row.get("标签")),
        "备注": normalize(row.get("备注")),
    }


def main() -> int:
    excel_path = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else DEFAULT_EXCEL
    output_path = Path(sys.argv[2]).expanduser() if len(sys.argv) > 2 else DEFAULT_OUTPUT

    if not excel_path.exists():
        print(f"Excel file not found: {excel_path}", file=sys.stderr)
        return 1

    workbook = load_workbook(excel_path, data_only=True)
    if REGISTER_SHEET not in workbook.sheetnames:
        payload = {
            "generatedAt": date.today().isoformat(),
            "sourceExcel": str(excel_path),
            "items": [],
        }
    else:
        sheet = workbook[REGISTER_SHEET]
        headers = [normalize(cell.value) for cell in sheet[1]]
        items = []

        for values in sheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values))
            skill_id = normalize(row.get("Skill_ID"))
            if not skill_id or skill_id in {"填写说明", "示例"}:
                continue
            item = row_to_item(row)
            if not item["L2 Skill名称"]:
                continue
            items.append(item)

        payload = {
            "generatedAt": date.today().isoformat(),
            "sourceExcel": str(excel_path),
            "items": items,
        }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(payload['items'])} team skill rows to {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
