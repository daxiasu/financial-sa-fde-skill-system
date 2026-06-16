#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""每日全量量化研究流程 v1.0
整合 research_cn + news_analyzer + quant_wiki
数据流：crawl_unified.py → 研报/新闻 → Excel报告
用法:
  python run_full_daily.py
  python run_full_daily.py --stocks 600519 000858 --funds 110022 001417
  python run_full_daily.py --output D:\输出目录
"""
from __future__ import annotations
import sys, json, time, sqlite3
from pathlib import Path
from datetime import datetime, date

SKILL_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(SKILL_DIR / "pkg"))

import subprocess

# ── 路径配置 ───────────────────────────────────────────────
OUTPUT_DIR = Path(r"C:\Users\11736\Desktop")
DESKTOP_OUTPUT = OUTPUT_DIR / "量化日报"

def _ensure_dirs():
    DESKTOP_OUTPUT.mkdir(parents=True, exist_ok=True)
    (SKILL_DIR / "pkg" / "data").mkdir(parents=True, exist_ok=True)

# ── 1. 数据抓取 ─────────────────────────────────────────────
def crawl_all_data(stocks: list, funds: list) -> dict:
    """从多个数据源抓取股票/基金/宏观数据"""
    from research_cn import (
        fetch_research_overview, fetch_all_macro,
        fetch_eastmoney_stock_news, fetch_ths_stock_news,
        fetch_xueqiu_stock_news
    )
    from fund_analyzer import fetch_fund_nav, fetch_fund_info
    from news_analyzer import analyze_news_batch

    data = {
        "macro": [],
        "research": {},      # code -> [items]
        "news": {},          # code -> [news]
        "funds": {},         # code -> info+nav
        "market_news": []
    }

    # 宏观数据
    print("  [1] 宏观数据...")
    try:
        data["macro"] = fetch_all_macro()
        print(f"      {len(data['macro'])} 项宏观指标")
    except Exception as e:
        print(f"      宏观数据异常: {e}")

    # 个股研究与新闻
    for code in (stocks or [])[:5]:
        code = code.strip().zfill(6)
        print(f"  [1] 研究 {code}...")
        try:
            # 研究资讯
            overview = fetch_research_overview(code=code)
            data["research"][code] = overview
            print(f"      研报 {overview.get('total_items',0)} 条")
        except Exception as e:
            print(f"      研究抓取异常: {e}")

        try:
            # 新闻
            news_list = []
            em = fetch_eastmoney_stock_news(code)
            ths = fetch_ths_stock_news(code)
            xq = fetch_xueqiu_stock_news(code)
            news_list.extend(em + ths + xq)
            data["news"][code] = news_list
            print(f"      新闻 {len(news_list)} 条")
        except Exception as e:
            print(f"      新闻抓取异常: {e}")

        time.sleep(0.5)

    # 基金数据
    for code in (funds or [])[:3]:
        code = code.strip().zfill(6)
        print(f"  [1] 基金 {code}...")
        try:
            nav = fetch_fund_nav(code)
            info = fetch_fund_info(code)
            data["funds"][code] = {"nav": nav, "info": info}
            print(f"      {nav.get('name','N/A')} 净值{nav.get('nav','N/A')}")
        except Exception as e:
            print(f"      基金数据异常: {e}")
        time.sleep(0.5)

    # 整体市场新闻（汇总分析）
    print("  [1] 市场情绪新闻...")
    try:
        # 抓取A股市场整体新闻
        market_news = fetch_eastmoney_stock_news("A股")
        if market_news:
            for n in market_news:
                n.setdefault("content", "")
            sentiment = analyze_news_batch(market_news[:50])
            data["market_news"] = {
                "total": sentiment["total"],
                "sentiment": sentiment["overall_sentiment"],
                "avg_score": sentiment["sentiment_avg"],
                "impact": sentiment["impact_counts"],
                "top_positive": sentiment.get("top_positive", [])[:3],
                "top_negative": sentiment.get("top_negative", [])[:3],
                "tag_counts": dict(list(sentiment["tag_counts"].items())[:8]),
            }
            print(f"      市场新闻{sentiment['total']}条 → {sentiment['overall_sentiment']}")
    except Exception as e:
        print(f"      市场新闻异常: {e}")

    return data

# ── 2. 生成 Excel 报告 ───────────────────────────────────────
def generate_excel_report(data: dict, report_date: str) -> Path:
    """生成 Excel 格式的每日量化报告"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [2] openpyxl 未安装，跳过Excel报告")
        return None

    print("  [2] 生成Excel报告...")
    wb = openpyxl.Workbook()

    # 样式定义
    DARK_BLUE  = "1F4E79"
    LIGHT_BLUE = "D9E1F2"
    MED_BLUE   = "BDD7EE"
    BLUE_FONT  = "0000FF"
    GREEN_FONT = "008000"
    GREY_FILL  = "F2F2F2"
    THIN = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def hdr_cell(ws, row, col, text, col_span=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_span > 1:
            for i in range(1, col_span):
                ws.cell(row=row, column=col+i).fill = PatternFill("solid", fgColor=DARK_BLUE)
        return cell

    def data_cell(ws, row, col, value, fmt=None, bold=False, bg=None, color=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = BORDER
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        if color:
            cell.font = Font(color=color, bold=bold)
        elif bold:
            cell.font = Font(bold=True)
        if fmt:
            cell.number_format = fmt
        cell.alignment = Alignment(horizontal="left", vertical="center")
        return cell

    # ── Sheet 1: 市场概览 ───────────────────────────────────
    ws1 = wb.active
    ws1.title = "市场概览"
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 16

    hdr_cell(ws1, 1, 1, f"每日量化报告 {report_date}", col_span=4)
    ws1.row_dimensions[1].height = 28

    # 宏观数据
    hdr_cell(ws1, 3, 1, "宏观指标", col_span=4)
    ws1.cell(row=3, column=1).alignment = Alignment(horizontal="left")
    for i, m in enumerate(data.get("macro", [])[:12], start=4):
        data_cell(ws1, i, 1, m.get("indicator",""))
        data_cell(ws1, i, 2, m.get("value","N/A"))
        data_cell(ws1, i, 3, m.get("unit",""))
        data_cell(ws1, i, 4, m.get("period",""))

    # 市场情绪
    r = len(data.get("macro", [])) + 6
    hdr_cell(ws1, r, 1, "市场情绪", col_span=4)
    ms = data.get("market_news", {})
    data_cell(ws1, r+1, 1, "新闻总数")
    data_cell(ws1, r+1, 2, ms.get("total", 0))
    data_cell(ws1, r+1, 3, "综合情绪")
    data_cell(ws1, r+1, 4, ms.get("sentiment","N/A"))
    data_cell(ws1, r+2, 1, "平均情绪分")
    data_cell(ws1, r+2, 2, f"{ms.get('avg_score',0):+.2f}", bg=MED_BLUE)
    ic = ms.get("impact", {})
    data_cell(ws1, r+3, 1, "利好/中性/利空")
    data_cell(ws1, r+3, 2, f"{ic.get('利好',0)}/{ic.get('中性',0)}/{ic.get('利空',0)}")

    # ── Sheet 2: 个股分析 ──────────────────────────────────
    ws2 = wb.create_sheet("个股分析")
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 12
    ws2.column_dimensions["G"].width = 20

    hdr_cell(ws2, 1, 1, "个股研究资讯摘要", col_span=7)
    headers = ["代码","数据源","标题/摘要","情绪","影响","标签","日期"]
    for c, h in enumerate(headers, 1):
        hdr_cell(ws2, 2, c, h)

    row = 3
    for code, overview in data.get("research", {}).items():
        for src, items in [("雪球","xueqiu"),("东财","eastmoney"),("同花顺","ths")]:
            src_items = overview.get(src, [])
            for item in src_items[:5]:
                title = (item.get("title") or item.get("text","") or "")[:40]
                if not title:
                    continue
                data_cell(ws2, row, 1, code, bold=True, bg=LIGHT_BLUE)
                data_cell(ws2, row, 2, src)
                data_cell(ws2, row, 3, title)
                data_cell(ws2, row, 4, item.get("sentiment","中性"))
                data_cell(ws2, row, 5, item.get("impact","中性"))
                data_cell(ws2, row, 6, ",".join(item.get("tags",[])))
                data_cell(ws2, row, 7, item.get("pub_date",""))
                row += 1
            if not src_items:
                data_cell(ws2, row, 1, code, bold=True, bg=LIGHT_BLUE)
                data_cell(ws2, row, 2, src)
                data_cell(ws2, row, 3, "（无数据）")
                row += 1

    # ── Sheet 3: 基金分析 ──────────────────────────────────
    ws3 = wb.create_sheet("基金分析")
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 12
    ws3.column_dimensions["E"].width = 12
    ws3.column_dimensions["F"].width = 12
    ws3.column_dimensions["G"].width = 12

    hdr_cell(ws3, 1, 1, "基金数据摘要", col_span=7)
    headers3 = ["代码","名称","类型","最新净值","估算涨跌%","基金经理","规模"]
    for c, h in enumerate(headers3, 1):
        hdr_cell(ws3, 2, c, h)

    row3 = 3
    for code, fdata in data.get("funds", {}).items():
        nav = fdata.get("nav", {})
        info = fdata.get("info", {})
        data_cell(ws3, row3, 1, code, bold=True, bg=LIGHT_BLUE)
        data_cell(ws3, row3, 2, nav.get("name", info.get("name","")))
        data_cell(ws3, row3, 3, info.get("type",""))
        data_cell(ws3, row3, 4, nav.get("nav","N/A"), fmt="0.0000")
        data_cell(ws3, row3, 5, nav.get("est_change_pct","N/A"), fmt="0.00%")
        data_cell(ws3, row3, 6, info.get("manager",""))
        data_cell(ws3, row3, 7, info.get("scale",""))
        row3 += 1

    # ── Sheet 4: 热点新闻 ─────────────────────────────────
    ws4 = wb.create_sheet("热点新闻")
    ws4.column_dimensions["A"].width = 10
    ws4.column_dimensions["B"].width = 40
    ws4.column_dimensions["C"].width = 10
    ws4.column_dimensions["D"].width = 12
    ws4.column_dimensions["E"].width = 20

    hdr_cell(ws4, 1, 1, "市场热点新闻", col_span=5)
    headers4 = ["代码","新闻摘要","情绪","影响","标签"]
    for c, h in enumerate(headers4, 1):
        hdr_cell(ws4, 2, c, h)

    row4 = 3
    news_all = []
    for code, news_list in data.get("news", {}).items():
        for n in news_list:
            n["_code"] = code
            news_all.append(n)

    for n in news_all[:50]:
        title = (n.get("title") or n.get("text","") or "")[:60]
        data_cell(ws4, row4, 1, n.get("_code",""), bg=LIGHT_BLUE)
        data_cell(ws4, row4, 2, title)
        data_cell(ws4, row4, 3, n.get("sentiment","中性"))
        data_cell(ws4, row4, 4, n.get("impact","中性"))
        data_cell(ws4, row4, 5, ",".join(n.get("tags",[])))
        row4 += 1

    # 保存
    out_path = DESKTOP_OUTPUT / f"量化日报_{report_date.replace('-','')}.xlsx"
    wb.save(out_path)
    print(f"      Excel已保存: {out_path}")
    return out_path

# ── 3. 保存 JSON 日报 ────────────────────────────────────────
def save_json_report(data: dict, report_date: str) -> Path:
    print("  [3] 保存JSON报告...")
    out_path = data_dir / f"daily_quant_{report_date.replace('-','')}.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump({
            "report_date": report_date,
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            **data
        }, f, ensure_ascii=False, indent=2)
    print(f"      JSON已保存: {out_path}")
    return out_path

# ── 4. 打印摘要 ─────────────────────────────────────────────
def print_summary(data: dict, report_date: str):
    print(f"\n{'='*65}")
    print(f"  每日量化研究简报 {report_date}")
    print(f"{'='*65}")

    # 宏观
    macro = data.get("macro", [])
    if macro:
        print(f"  宏观: {len(macro)} 项指标")
        for m in macro[:4]:
            print(f"    {m.get('indicator','?')}: {m.get('value','?')} {m.get('unit','?')}")

    # 市场情绪
    ms = data.get("market_news", {})
    if ms:
        print(f"  市场情绪: {ms.get('sentiment','N/A')} (avg={ms.get('avg_score',0):+.2f})")
        ic = ms.get("impact", {})
        print(f"    利好{ic.get('利好',0)}条 | 中性{ic.get('中性',0)}条 | 利空{ic.get('利空',0)}条")

    # 个股
    research = data.get("research", {})
    if research:
        print(f"  个股研究: {len(research)} 只股票有数据")
        for code, overview in list(research.items())[:3]:
            total = overview.get("total_items", 0)
            print(f"    {code}: {total} 条研报/资讯")

    # 基金
    funds = data.get("funds", {})
    if funds:
        print(f"  基金: {len(funds)} 只")
        for code, fdata in funds.items():
            nav = fdata.get("nav", {})
            print(f"    {code}: {nav.get('name','?')} 净值={nav.get('nav','?')}")

    print(f"{'='*65}\n")

# ── 主流程 ───────────────────────────────────────────────────
def run_full_daily(stocks: list = None, funds: list = None, output_dir: str = None):
    """每日全量研究流程"""
    global DESKTOP_OUTPUT
    if output_dir:
        DESKTOP_OUTPUT = Path(output_dir)

    _ensure_dirs()
    report_date = date.today().strftime("%Y-%m-%d")

    print(f"\n{'='*65}")
    print(f"  每日量化研究全流程 v1.0")
    print(f"  时间: {report_date} {datetime.now().strftime('%H:%M:%S')}")
    print(f"  输出: {DESKTOP_OUTPUT}")
    print(f"{'='*65}")

    # Step 1: 抓取数据
    print("\n[Step 1] 数据抓取")
    data = crawl_all_data(stocks or [], funds or [])

    # Step 2: Excel报告
    print("\n[Step 2] 生成Excel报告")
    excel_path = generate_excel_report(data, report_date)

    # Step 3: JSON日报
    print("\n[Step 3] 保存JSON日报")
    json_path = save_json_report(data, report_date)

    # Step 4: 打印摘要
    print("\n[Step 4] 报告摘要")
    print_summary(data, report_date)

    print(f"✅ 全流程完成！")
    print(f"   Excel: {excel_path}")
    print(f"   JSON:  {json_path}")
    return data

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="每日量化研究全流程")
    parser.add_argument("--stocks", nargs="*", default=[], help="股票代码列表")
    parser.add_argument("--funds", nargs="*", default=[], help="基金代码列表")
    parser.add_argument("--output", help="输出目录")
    args = parser.parse_args()

    run_full_daily(
        stocks=args.stocks or ["600519", "000858", "300750"],
        funds=args.funds or ["110022", "001417"],
        output_dir=args.output,
    )
